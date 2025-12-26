#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
請求書処理完全自動化システム v5

改修内容:
- PDFファイル名からアンダーバー削除
- 会社マスターExcelを2シート構造に対応（会社マスタ/メール）
- CC機能追加
- 保存先を会社マスター「保存先」シートから自動取得
- 年月フォルダ自動生成（YYYY年MM月形式）
- 同月フォルダが既存の場合は上書き保存

Author: Hiroshima Plastic
Version: 5.0.0
"""

from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from pathlib import Path
import os
import sys
import re
import tempfile
import logging
from datetime import datetime

# PDF処理
from pypdf import PdfWriter, PdfReader
import pdfplumber

# Excel処理
import openpyxl

# PDF生成
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

# GUI
import tkinter as tk
from tkinter import filedialog, messagebox

# Outlook連携
try:
    import win32com.client
    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False


# =====================================
# ロギング設定
# =====================================
logging.basicConfig(
    level=logging.INFO,  # INFOレベルに戻す
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('invoice_automation.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# pdfplumberのログを抑制（DEBUGログが大量に出るため）
logging.getLogger('pdfminer').setLevel(logging.WARNING)
logging.getLogger('pdfplumber').setLevel(logging.WARNING)


# =====================================
# 定数定義
# =====================================
@dataclass(frozen=True)
class SealConfig:
    """電子印設定"""
    # 印鑑サイズ
    PERSONAL_SIZE: Tuple[int, int] = (20, 20)  # 個人印（管理者・担当者）
    COMPANY_SIZE: Tuple[int, int] = (60, 60)   # 社印
    
    # 押印位置（座標）
    POSITIONS: Dict[str, Tuple[int, int]] = None
    
    # 押印対象ページ
    FIRST_PAGE_ONLY: bool = True
    
    def __post_init__(self):
        """イミュータブルな辞書を設定"""
        if self.POSITIONS is None:
            object.__setattr__(self, 'POSITIONS', {
                '管理者': (457, 655),
                '担当者': (498, 655),
                '社印': (500, 680),
            })


@dataclass(frozen=True)
class FileConfig:
    """ファイル名設定"""
    # ファイル名形式: YYMMDDCompanyName請求書.pdf
    FILENAME_PATTERN: str = "{date}{company}請求書.pdf"
    INVALID_CHARS: str = r'[\\/:*?"<>|]'


@dataclass
class CompanyInfo:
    """会社情報"""
    name: str
    email: str
    cc: Optional[str] = None


@dataclass
class EmailTemplate:
    """メールテンプレート"""
    subject: str
    body: str


@dataclass
class InvoiceInfo:
    """請求書情報"""
    invoice_number: str
    company: str
    pages: List[int]
    close_date_full: Optional[str] = None  # YYYY-MM-DD
    close_date_short: Optional[str] = None  # YYMMDD
    is_copy: bool = False
    no_transaction: bool = False
    is_blank: bool = False


# =====================================
# Excel読み込みクラス
# =====================================
class CompanyMasterReader:
    """会社マスターExcel読み込みクラス"""
    
    def __init__(self, excel_path: str):
        """
        Args:
            excel_path: 会社マスターExcelのパス
        """
        self.excel_path = Path(excel_path)
        self.companies: Dict[str, CompanyInfo] = {}
        self.email_template: Optional[EmailTemplate] = None
        self.output_base_path: Optional[Path] = None  # 保存先ベースパス
        
    def load(self) -> bool:
        """
        会社マスターを読み込み
        
        Returns:
            bool: 読み込み成功時True
        """
        logger.info(f"会社マスター読み込み開始: {self.excel_path}")
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            
            # シート1: 会社マスタ
            if '会社マスタ' not in wb.sheetnames:
                logger.error("シート '会社マスタ' が見つかりません")
                return False
            
            # シート2: メール
            if 'メール' not in wb.sheetnames:
                logger.error("シート 'メール' が見つかりません")
                return False
            
            # シート3: 保存先（オプション）
            has_output_sheet = '保存先' in wb.sheetnames
            
            # 会社情報を読み込み
            if not self._load_companies(wb['会社マスタ']):
                return False
            
            # メールテンプレートを読み込み
            if not self._load_email_template(wb['メール']):
                return False
            
            # 保存先パスを読み込み（オプション）
            if has_output_sheet:
                self._load_output_path(wb['保存先'])
            
            logger.info(f"✓ {len(self.companies)}社の情報を読み込みました")
            self._log_loaded_companies()
            
            return True
            
        except Exception as e:
            logger.error(f"会社マスター読み込み失敗: {e}", exc_info=True)
            return False
    
    def _load_companies(self, ws) -> bool:
        """
        会社マスタシートから会社情報を読み込み
        
        Args:
            ws: 会社マスタワークシート
            
        Returns:
            bool: 読み込み成功時True
        """
        try:
            # ヘッダー行をスキップ（2行目から）
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            if not rows:
                logger.error("会社マスタシートが空です")
                return False
            
            for row in rows:
                # A列（会社名）が空ならスキップ
                if not row[0]:
                    continue
                
                company_name = row[0].strip()
                email = row[1].strip() if len(row) > 1 and row[1] else None
                cc = row[2].strip() if len(row) > 2 and row[2] else None
                
                if not email:
                    logger.warning(f"会社 '{company_name}' のメールアドレスが未設定です")
                
                self.companies[company_name] = CompanyInfo(
                    name=company_name,
                    email=email,
                    cc=cc
                )
            
            return len(self.companies) > 0
            
        except Exception as e:
            logger.error(f"会社情報読み込みエラー: {e}", exc_info=True)
            return False
    
    def _load_email_template(self, ws) -> bool:
        """
        メールシートからテンプレートを読み込み
        
        シート構造:
        A1: メールタイトル (ヘッダー)
        A2: 【your conpany name】YYYY年MM月分 ご請求書送付の件 (実データ)
        A3: (空行)
        A4: メール本文 (ヘッダー)
        A5: A:A (実データ開始)
        A6: ご担当者様
        ...
        
        Args:
            ws: メールワークシート
            
        Returns:
            bool: 読み込み成功時True
        """
        try:
            # メールタイトルを取得 (A2)
            subject = ws['A2'].value
            
            # メール本文を取得
            # "メール本文" というヘッダーを探す
            body_start_row = None
            for row_idx in range(1, 20):
                cell_value = ws[f'A{row_idx}'].value
                if cell_value and "メール本文" in str(cell_value):
                    body_start_row = row_idx + 1  # ヘッダーの次の行
                    break
            
            if body_start_row is None:
                logger.error("「メール本文」ヘッダーが見つかりません")
                return False
            
            # 本文を複数行結合
            body_lines = []
            for row_idx in range(body_start_row, body_start_row + 50):
                cell_value = ws[f'A{row_idx}'].value
                if cell_value is None:
                    break  # 空行で終了
                body_lines.append(str(cell_value))
            
            body = '\n'.join(body_lines)
            
            if not subject or not body:
                logger.error("メールテンプレートが未設定です")
                logger.error(f"  件名が空: {not subject}")
                logger.error(f"  本文が空: {not body}")
                return False
            
            # 文字列に変換してstrip
            subject_str = str(subject).strip()
            body_str = body.strip()
            
            self.email_template = EmailTemplate(
                subject=subject_str,
                body=body_str
            )
            
            logger.info(f"✓ メールテンプレート読み込み完了")
            logger.info(f"  件名: {self.email_template.subject}")
            logger.info(f"  本文: {self.email_template.body[:50]}...")
            
            return True
            
        except Exception as e:
            logger.error(f"メールテンプレート読み込みエラー: {e}", exc_info=True)
            return False
    
    def _log_loaded_companies(self):
        """読み込んだ会社情報をログ出力"""
        logger.info("登録された会社一覧:")
        for company, info in self.companies.items():
            logger.info(f"  - {company}")
            logger.info(f"      メール: {info.email or '(未設定)'}")
            if info.cc:
                logger.info(f"      CC: {info.cc}")
    
    def _load_output_path(self, ws) -> bool:
        """
        保存先シートからベースパスを読み込み
        
        シート構造:
        A1: フォルダ
        B1: \\Hiropula-srv\共通\請求書
        
        Args:
            ws: 保存先ワークシート
            
        Returns:
            bool: 読み込み成功時True
        """
        try:
            # B1セルから保存先パスを取得
            output_path = ws['B1'].value
            
            if output_path:
                self.output_base_path = Path(output_path)
                logger.info(f"✓ 保存先パス設定: {self.output_base_path}")
                return True
            else:
                logger.warning("保存先パスが未設定です（B1セルが空）")
                return False
                
        except Exception as e:
            logger.error(f"保存先パス読み込みエラー: {e}", exc_info=True)
            return False
    
    def get_company_info(self, company_name: str) -> Optional[CompanyInfo]:
        """
        会社情報を取得
        
        Args:
            company_name: 会社名
            
        Returns:
            CompanyInfo: 会社情報（存在しない場合None）
        """
        return self.companies.get(company_name)
    
    def get_email_for_company(self, company_name: str) -> Tuple[str, str]:
        """
        会社用のメール件名・本文を取得
        
        Args:
            company_name: 会社名
            
        Returns:
            Tuple[str, str]: (件名, 本文)
        """
        if not self.email_template:
            return "", ""
        
        # A:A を会社名に置換
        subject = self.email_template.subject
        body = self.email_template.body.replace('A:A', company_name)
        
        return subject, body


# =====================================
# 電子印管理クラス
# =====================================
class SealManager:
    """電子印画像管理クラス"""
    
    def __init__(self, seal_dir: str):
        """
        Args:
            seal_dir: 電子印画像フォルダパス
        """
        self.seal_dir = Path(seal_dir)
        self.seal_images: Dict[str, Path] = {}
        self.config = SealConfig()
        
    def load(self) -> bool:
        """
        電子印画像を読み込み
        
        Returns:
            bool: 読み込み成功時True
        """
        logger.info(f"電子印画像読み込み開始: {self.seal_dir}")
        
        if not self.seal_dir.exists():
            logger.error(f"電子印フォルダが見つかりません: {self.seal_dir}")
            return False
        
        try:
            for seal_file in self.seal_dir.glob('*.png'):
                seal_name = seal_file.name
                self.seal_images[seal_name] = seal_file
                logger.info(f"  ✓ {seal_name}")
            
            logger.info(f"✓ {len(self.seal_images)}個の電子印を読み込みました")
            return len(self.seal_images) > 0
            
        except Exception as e:
            logger.error(f"電子印読み込みエラー: {e}", exc_info=True)
            return False
    
    def get_seal_path(self, seal_name: str) -> Optional[Path]:
        """
        電子印画像パスを取得
        
        Args:
            seal_name: 印鑑名（例: '管理者.png'）
            
        Returns:
            Path: 画像パス（存在しない場合None）
        """
        return self.seal_images.get(seal_name)
    
    def has_required_seals(self) -> bool:
        """
        必須印鑑（管理者・担当者）が揃っているか確認
        
        Returns:
            bool: 揃っている場合True
        """
        return '管理者.png' in self.seal_images and '担当者.png' in self.seal_images


# =====================================
# PDF処理クラス
# =====================================
class InvoicePDFProcessor:
    """請求書PDF処理クラス"""
    
    def __init__(self, seal_manager: SealManager):
        """
        Args:
            seal_manager: 電子印管理インスタンス
        """
        self.seal_manager = seal_manager
        self.invoices: List[InvoiceInfo] = []
        
    def split_pdf(self, input_pdf: Path) -> bool:
        """
        PDFを会社ごとに分割
        
        Args:
            input_pdf: 入力PDFパス
            
        Returns:
            bool: 分割成功時True
        """
        logger.info("="*70)
        logger.info(f"PDF分割開始: {input_pdf.name}")
        logger.info("="*70)
        
        try:
            with pdfplumber.open(input_pdf) as pdf:
                total_pages = len(pdf.pages)
                logger.info(f"総ページ数: {total_pages}")
                
                current_invoice = None
                
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    invoice_info = self._extract_invoice_info(text, page_num)
                    
                    if not invoice_info:
                        continue
                    
                    # スキップ条件
                    if invoice_info.is_copy:
                        logger.info(f"ページ{page_num}: {invoice_info.invoice_number} - {invoice_info.company} [控え] ← スキップ")
                        continue
                    
                    if invoice_info.no_transaction:
                        logger.info(f"ページ{page_num}: {invoice_info.invoice_number} - {invoice_info.company} [当月取引なし] ← スキップ")
                        continue
                    
                    if invoice_info.is_blank:
                        logger.info(f"ページ{page_num}: [ほぼ空白ページ] ← スキップ")
                        continue
                    
                    logger.info(f"ページ{page_num}: {invoice_info.invoice_number} - {invoice_info.company}")
                    
                    # 請求書番号の基本部分を取得
                    base_number = invoice_info.invoice_number.split('-')[0]
                    current_base = current_invoice.invoice_number.split('-')[0] if current_invoice else None
                    
                    if current_invoice is None or current_base != base_number:
                        # 新しい請求書
                        if current_invoice:
                            self.invoices.append(current_invoice)
                        current_invoice = invoice_info
                    else:
                        # 既存請求書にページ追加
                        current_invoice.pages.append(page_num)
                
                # 最後の請求書を追加
                if current_invoice:
                    self.invoices.append(current_invoice)
                
                logger.info(f"✓ {len(self.invoices)}社の請求書を検出")
                return True
                
        except Exception as e:
            logger.error(f"PDF分割エラー: {e}", exc_info=True)
            return False
    
    def _extract_invoice_info(self, text: str, page_num: int) -> Optional[InvoiceInfo]:
        """
        請求書情報を抽出
        
        Args:
            text: ページテキスト
            page_num: ページ番号
            
        Returns:
            InvoiceInfo: 請求書情報（抽出失敗時None）
        """
        if not text:
            return None
        
        # 請求書番号を抽出
        match = re.search(r'№\s*(\d+(?:-\d+)?)', text)
        if not match:
            return None
        
        invoice_number = match.group(1)
        
        # 各種フラグ判定
        is_copy = '（控）' in text or '(控)' in text
        no_transaction = '当月のお取引はございません' in text or '当月のお取引' in text
        is_blank = len(text.strip()) < 100
        
        # 会社名を抽出
        company_name = self._extract_company_name(text)
        
        # 締め日を抽出
        close_date_full, close_date_short = self._extract_close_date(text)
        
        return InvoiceInfo(
            invoice_number=invoice_number,
            company=company_name,
            pages=[page_num],
            close_date_full=close_date_full,
            close_date_short=close_date_short,
            is_copy=is_copy,
            no_transaction=no_transaction,
            is_blank=is_blank
        )
    
    def _extract_company_name(self, text: str) -> str:
        """
        会社名を抽出
        
        Args:
            text: ページテキスト
            
        Returns:
            str: 会社名
        """
        match = re.search(r'([^\n]+?)\s*御中', text)
        if match:
            company_name = match.group(1).strip()
            # 郵便番号や住所を除去
            company_name = re.sub(r'〒.*?\n', '', company_name)
            company_name = re.sub(r'[都道府県].*', '', company_name, flags=re.DOTALL)
            return company_name.split('\n')[-1].strip()
        return '不明'
    
    def _extract_close_date(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """
        締め日を抽出
        
        Args:
            text: ページテキスト
            
        Returns:
            Tuple[Optional[str], Optional[str]]: (YYYY-MM-DD形式, YYMMDD形式)
        """
        match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日締切分', text)
        if match:
            year = match.group(1)
            month = int(match.group(2))
            day = int(match.group(3))
            
            close_date_full = f"{year}-{month:02d}-{day:02d}"
            close_date_short = f"{year[2:]}{month:02d}{day:02d}"
            
            return close_date_full, close_date_short
        
        return None, None
    
    def create_pdf_with_seal(
        self,
        input_pdf: Path,
        invoice: InvoiceInfo,
        output_dir: Path
    ) -> Optional[Path]:
        """
        分割PDFを作成し、1ページ目に電子印を押印
        
        Args:
            input_pdf: 入力PDFパス
            invoice: 請求書情報
            output_dir: 出力先ディレクトリ
            
        Returns:
            Path: 作成されたPDFパス（失敗時None）
        """
        try:
            # ファイル名生成（アンダーバー除去）
            filename = self._generate_filename(invoice)
            
            temp_pdf = output_dir / f"temp_{filename}"
            final_pdf = output_dir / filename
            
            # ページ抽出
            reader = PdfReader(input_pdf)
            writer = PdfWriter()
            
            for page_num in invoice.pages:
                writer.add_page(reader.pages[page_num - 1])
            
            with open(temp_pdf, 'wb') as f:
                writer.write(f)
            
            # 電子印押印
            if self._add_seals_to_pdf(temp_pdf, final_pdf):
                temp_pdf.unlink()
            else:
                temp_pdf.rename(final_pdf)
            
            return final_pdf
            
        except Exception as e:
            logger.error(f"PDF作成エラー: {e}", exc_info=True)
            return None
    
    def _generate_filename(self, invoice: InvoiceInfo) -> str:
        """
        PDFファイル名を生成（アンダーバー除去）
        
        Args:
            invoice: 請求書情報
            
        Returns:
            str: ファイル名
        """
        # 会社名から無効な文字を除去
        safe_company = re.sub(FileConfig.INVALID_CHARS, '_', invoice.company)
        
        # YYMMDDCompanyName請求書.pdf
        if invoice.close_date_short:
            return f"{invoice.close_date_short}{safe_company}請求書.pdf"
        else:
            return f"{safe_company}請求書.pdf"
    
    def _add_seals_to_pdf(self, input_pdf: Path, output_pdf: Path) -> bool:
        """
        PDFの1ページ目に電子印を押印
        
        Args:
            input_pdf: 入力PDFパス
            output_pdf: 出力PDFパス
            
        Returns:
            bool: 成功時True
        """
        logger.info("  電子印押印中... (1ページ目のみ)")
        
        try:
            reader = PdfReader(input_pdf)
            writer = PdfWriter()
            
            # 印鑑パス取得
            kanrisha_seal = self.seal_manager.get_seal_path('管理者.png')
            tantousha_seal = self.seal_manager.get_seal_path('担当者.png')
            company_seal = self.seal_manager.get_seal_path('社印.png')
            
            # デバッグ情報
            config = self.seal_manager.config
            logger.debug(f"  印鑑ファイル:")
            logger.debug(f"    管理者.png: {kanrisha_seal}")
            logger.debug(f"    担当者.png: {tantousha_seal}")
            logger.debug(f"    社印.png: {company_seal}")
            
            # 必須印鑑チェック
            if not kanrisha_seal or not tantousha_seal:
                logger.warning("管理者.png または 担当者.png が見つかりません")
                # 印鑑なしでコピー
                for page in reader.pages:
                    writer.add_page(page)
                with open(output_pdf, 'wb') as f:
                    writer.write(f)
                return True
            
            # 各ページを処理
            temp_files = []
            
            for page_num, page in enumerate(reader.pages):
                if page_num == 0:
                    # 1ページ目のみ押印
                    logger.debug("  1ページ目に印鑑を押印中...")
                    
                    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
                        tmp_pdf_path = tmp_file.name
                        temp_files.append(tmp_pdf_path)
                        
                        # 印鑑レイヤーPDF作成
                        c = canvas.Canvas(tmp_pdf_path, pagesize=A4)
                        
                        # 管理者印
                        c.drawImage(
                            str(kanrisha_seal),
                            config.POSITIONS['管理者'][0],
                            config.POSITIONS['管理者'][1],
                            width=config.PERSONAL_SIZE[0],
                            height=config.PERSONAL_SIZE[1],
                            mask='auto',
                            preserveAspectRatio=True
                        )
                        
                        # 担当者印
                        c.drawImage(
                            str(tantousha_seal),
                            config.POSITIONS['担当者'][0],
                            config.POSITIONS['担当者'][1],
                            width=config.PERSONAL_SIZE[0],
                            height=config.PERSONAL_SIZE[1],
                            mask='auto',
                            preserveAspectRatio=True
                        )
                        
                        # 社印（任意）
                        if company_seal and company_seal.exists():
                            c.drawImage(
                                str(company_seal),
                                config.POSITIONS['社印'][0],
                                config.POSITIONS['社印'][1],
                                width=config.COMPANY_SIZE[0],
                                height=config.COMPANY_SIZE[1],
                                mask='auto',
                                preserveAspectRatio=True
                            )
                        
                        c.save()
                    
                    # ページマージ
                    seal_pdf = PdfReader(tmp_pdf_path)
                    page.merge_page(seal_pdf.pages[0])
                else:
                    logger.debug(f"  {page_num + 1}ページ目: 印鑑なし（スキップ）")
                
                writer.add_page(page)
            
            # PDF保存
            with open(output_pdf, 'wb') as f:
                writer.write(f)
            
            # 一時ファイル削除
            for tmp_file in temp_files:
                try:
                    os.unlink(tmp_file)
                except Exception:
                    pass
            
            logger.info("  ✓ 電子印押印完了（管理者・担当者・社印）")
            return True
            
        except Exception as e:
            logger.error(f"電子印押印エラー: {e}", exc_info=True)
            return False


# =====================================
# Outlookメール作成クラス
# =====================================
class OutlookMailCreator:
    """Outlookメール下書き作成クラス"""
    
    def __init__(self, company_master: CompanyMasterReader):
        """
        Args:
            company_master: 会社マスターリーダー
        """
        self.company_master = company_master
        
    def create_draft(
        self,
        company_name: str,
        pdf_path: Path,
        close_date_full: Optional[str] = None
    ) -> bool:
        """
        Outlookメール下書きを作成
        
        Args:
            company_name: 会社名
            pdf_path: 添付PDFパス
            close_date_full: 締め日（YYYY-MM-DD形式）
            
        Returns:
            bool: 成功時True
        """
        logger.info("  Outlookメール下書き作成中...")
        
        if not OUTLOOK_AVAILABLE:
            logger.warning("Outlook連携スキップ（pywin32未インストール）")
            return False
        
        # 会社情報取得
        company_info = self.company_master.get_company_info(company_name)
        if not company_info:
            logger.warning(f"会社 '{company_name}' のメール情報が見つかりません")
            return False
        
        # メールテンプレート取得
        subject, body = self.company_master.get_email_for_company(company_name)
        
        # 件名の日付置換
        subject = self._replace_date_placeholder(subject, close_date_full)
        
        logger.debug(f"    宛先: {company_info.email}")
        if company_info.cc:
            logger.debug(f"    CC: {company_info.cc}")
        logger.debug(f"    件名: {subject}")
        logger.debug(f"    添付: {pdf_path.name}")
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            
            mail.To = company_info.email
            if company_info.cc:
                mail.CC = company_info.cc
            mail.Subject = subject
            mail.Body = body
            mail.Attachments.Add(str(pdf_path.absolute()))
            
            mail.Save()
            
            logger.info("  ✓ Outlookメール下書き作成完了")
            return True
            
        except Exception as e:
            logger.error(f"Outlookメール作成エラー: {e}", exc_info=True)
            return False
    
    def _replace_date_placeholder(
        self,
        subject: str,
        close_date_full: Optional[str]
    ) -> str:
        """
        件名のYYYY年MM月を置換
        
        Args:
            subject: 件名テンプレート
            close_date_full: 締め日（YYYY-MM-DD形式）
            
        Returns:
            str: 置換後の件名
        """
        if not close_date_full or 'YYYY年MM月' not in subject:
            return subject
        
        # YYYY-MM-DD → YYYY年M月（月の0埋めなし）
        year, month, day = close_date_full.split('-')
        month_int = int(month)
        
        return subject.replace('YYYY年MM月', f'{year}年{month_int}月')


# =====================================
# メイン処理クラス
# =====================================
class InvoiceAutomationSystem:
    """請求書処理自動化システム メインクラス"""
    
    def __init__(
        self,
        company_master: CompanyMasterReader,
        seal_manager: SealManager
    ):
        """
        Args:
            company_master: 会社マスターリーダー
            seal_manager: 電子印管理インスタンス
        """
        self.company_master = company_master
        self.seal_manager = seal_manager
        self.pdf_processor = InvoicePDFProcessor(seal_manager)
        self.mail_creator = OutlookMailCreator(company_master)
        
    def process(
        self,
        input_pdf: Path,
        output_base_dir: Path
    ) -> List[Dict[str, Any]]:
        """
        請求書を一括処理
        
        Args:
            input_pdf: 入力PDFパス
            output_base_dir: 出力先ベースディレクトリ
            
        Returns:
            List[Dict]: 処理結果リスト
        """
        logger.info("="*70)
        logger.info("請求書一括処理開始")
        logger.info("="*70)
        
        results = []
        
        # 締め日から年月フォルダを判定（最初の請求書から取得）
        month_folder_name = None
        if self.pdf_processor.invoices:
            first_invoice = self.pdf_processor.invoices[0]
            if first_invoice.close_date_full:
                # YYYY-MM-DD → YYYY年MM月
                year, month, day = first_invoice.close_date_full.split('-')
                month_int = int(month)
                month_folder_name = f"{year}年{month_int}月"
        
        # 出力ディレクトリ決定
        if month_folder_name:
            output_dir = output_base_dir / month_folder_name
            logger.info(f"✓ 月別フォルダ: {month_folder_name}")
        else:
            output_dir = output_base_dir
            logger.warning("締め日が取得できないため、ベースフォルダに保存します")
        
        # ディレクトリ作成（既存の場合はスキップ）
        if output_dir.exists():
            logger.info(f"✓ 既存フォルダを使用: {output_dir}")
        else:
            output_dir.mkdir(parents=True, exist_ok=True)
            logger.info(f"✓ 新規フォルダ作成: {output_dir}")
        
        for idx, invoice in enumerate(self.pdf_processor.invoices, 1):
            logger.info(f"[{idx}/{len(self.pdf_processor.invoices)}] {invoice.company}")
            logger.info(f"  ページ: {invoice.pages}")
            if invoice.close_date_short:
                logger.info(f"  締め日: {invoice.close_date_short}")
            
            # PDF作成
            pdf_path = self.pdf_processor.create_pdf_with_seal(
                input_pdf,
                invoice,
                output_dir
            )
            
            if pdf_path:
                logger.info(f"  ✓ PDF作成完了: {pdf_path.name}")
                
                # Outlookメール作成
                if OUTLOOK_AVAILABLE:
                    self.mail_creator.create_draft(
                        invoice.company,
                        pdf_path,
                        invoice.close_date_full
                    )
                
                results.append({
                    'company': invoice.company,
                    'pdf': pdf_path,
                    'success': True
                })
            else:
                logger.error(f"  ✗ PDF作成失敗: {invoice.company}")
                results.append({
                    'company': invoice.company,
                    'success': False
                })
            
            logger.info("")
        
        return results


# =====================================
# ファイル検出ユーティリティ
# =====================================
def find_company_master(base_dir: Path) -> Optional[Path]:
    """
    会社マスター.xlsxを自動検出
    
    Args:
        base_dir: 基準ディレクトリ
        
    Returns:
        Path: 検出されたパス（見つからない場合None）
    """
    # 同じディレクトリ
    master_path = base_dir / "会社マスター.xlsx"
    if master_path.exists():
        return master_path
    
    # 下の階層（1階層のみ）
    for item in base_dir.iterdir():
        if item.is_dir():
            master_path = item / "会社マスター.xlsx"
            if master_path.exists():
                return master_path
    
    return None


def find_seal_directory(base_dir: Path) -> Optional[Path]:
    """
    電子印フォルダを自動検出
    
    Args:
        base_dir: 基準ディレクトリ
        
    Returns:
        Path: 検出されたパス（見つからない場合None）
    """
    # 同じディレクトリ
    seal_dir = base_dir / "電子印"
    if seal_dir.exists() and seal_dir.is_dir():
        return seal_dir
    
    # 下の階層（1階層のみ）
    for item in base_dir.iterdir():
        if item.is_dir():
            seal_dir = item / "電子印"
            if seal_dir.exists() and seal_dir.is_dir():
                return seal_dir
    
    return None


def get_output_directory(company_master: CompanyMasterReader) -> Optional[Path]:
    """
    保存先ディレクトリを取得
    
    会社マスターに「保存先」シートがあればそのパスを使用、
    なければダイアログで選択
    
    Args:
        company_master: 会社マスターリーダー
        
    Returns:
        Path: 保存先ベースパス（キャンセル時None）
    """
    # 会社マスターに保存先が設定されている場合
    if company_master.output_base_path:
        logger.info(f"✓ 会社マスターから保存先を取得: {company_master.output_base_path}")
        
        # パスの存在確認
        if not company_master.output_base_path.exists():
            logger.warning(f"指定された保存先が存在しません: {company_master.output_base_path}")
            logger.info("  手動で選択してください")
            return select_output_directory_manual()
        
        return company_master.output_base_path
    
    # 保存先が未設定の場合は手動選択
    logger.info("会社マスターに保存先が未設定です")
    logger.info("  手動で選択してください")
    return select_output_directory_manual()


def select_output_directory_manual() -> Optional[Path]:
    """
    保存先ディレクトリを手動選択
    
    Returns:
        Path: 選択されたパス（キャンセル時None）
    """
    root = tk.Tk()
    root.withdraw()
    
    # ダイアログで保存先を選択
    output_dir = filedialog.askdirectory(
        title="保存先フォルダを選択してください"
    )
    
    if not output_dir:
        return None
    
    return Path(output_dir)


# =====================================
# メイン処理
# =====================================
def main():
    """メイン処理エントリーポイント"""
    logger.info("="*70)
    logger.info("請求書処理完全自動化システム v5")
    logger.info("（1ページ目のみ・管理者/担当者/社印対応・CC対応）")
    logger.info("="*70)
    
    root = tk.Tk()
    root.withdraw()
    
    # ステップ1: PDFを選択
    input_pdf_path = filedialog.askopenfilename(
        title="請求書PDFを選択してください",
        filetypes=[("PDF files", "*.pdf")]
    )
    
    if not input_pdf_path:
        logger.info("処理をキャンセルしました")
        sys.exit(0)
    
    input_pdf = Path(input_pdf_path)
    base_dir = input_pdf.parent
    
    logger.info(f"選択されたPDF: {input_pdf.name}")
    logger.info(f"PDFの場所: {base_dir}")
    
    # ステップ2: 会社マスターを自動検出
    logger.info("会社マスターを検索中...")
    auto_master = find_company_master(base_dir)
    
    if auto_master:
        logger.info(f"✓ 自動検出: {auto_master.relative_to(base_dir)}")
        master_path = auto_master
    else:
        logger.info("✗ 自動検出できませんでした")
        logger.info("  手動で選択してください")
        master_path_str = filedialog.askopenfilename(
            title="会社マスターExcelを選択してください",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not master_path_str:
            logger.info("処理をキャンセルしました")
            sys.exit(0)
        
        master_path = Path(master_path_str)
    
    # ステップ3: 電子印フォルダを自動検出
    logger.info("電子印フォルダを検索中...")
    auto_seal_dir = find_seal_directory(base_dir)
    
    if auto_seal_dir:
        logger.info(f"✓ 自動検出: {auto_seal_dir.relative_to(base_dir)}")
        seal_dir = auto_seal_dir
    else:
        logger.info("✗ 自動検出できませんでした")
        logger.info("  手動で選択してください")
        seal_dir_str = filedialog.askdirectory(
            title="電子印画像フォルダを選択してください"
        )
        
        if not seal_dir_str:
            logger.info("処理をキャンセルしました")
            sys.exit(0)
        
        seal_dir = Path(seal_dir_str)
    
    # ステップ4: データ読み込み
    company_master = CompanyMasterReader(str(master_path))
    if not company_master.load():
        logger.error("会社マスター読み込みに失敗しました")
        sys.exit(1)
    
    seal_manager = SealManager(str(seal_dir))
    if not seal_manager.load():
        logger.error("電子印読み込みに失敗しました")
        sys.exit(1)
    
    # ステップ5: 保存先ベースディレクトリを取得
    logger.info("保存先フォルダを確認中...")
    output_base_dir = get_output_directory(company_master)
    
    if not output_base_dir:
        logger.info("処理をキャンセルしました")
        sys.exit(0)
    
    logger.info(f"✓ 保存先ベース: {output_base_dir}")
    
    # ステップ6: PDF分割
    system = InvoiceAutomationSystem(company_master, seal_manager)
    if not system.pdf_processor.split_pdf(input_pdf):
        logger.error("PDF分割に失敗しました")
        sys.exit(1)
    
    # ステップ7: 一括処理
    results = system.process(input_pdf, output_base_dir)
    
    # 最終的な出力先を取得（年月フォルダが作成されている場合）
    if results and 'pdf' in results[0]:
        final_output_dir = results[0]['pdf'].parent
    else:
        final_output_dir = output_base_dir
    
    # 結果表示
    logger.info("="*70)
    logger.info("処理完了")
    logger.info("="*70)
    logger.info(f"成功: {sum(1 for r in results if r['success'])}/{len(results)}社")
    logger.info(f"出力先: {final_output_dir}")
    
    messagebox.showinfo(
        "処理完了",
        f"請求書処理が完了しました\n\n"
        f"処理件数: {len(results)}社\n"
        f"成功: {sum(1 for r in results if r['success'])}社\n"
        f"出力先: {final_output_dir}"
    )


if __name__ == "__main__":
    try:
        # tkinterのsimpledialogをインポート（新規フォルダ作成用）
        import tkinter.simpledialog
        main()
    except Exception as e:
        logger.critical(f"予期しないエラー: {e}", exc_info=True)
        messagebox.showerror("エラー", f"システムエラーが発生しました:\n{e}")
        sys.exit(1)
